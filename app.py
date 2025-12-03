import os
import sqlite3
import random
from datetime import datetime
from io import BytesIO

from flask import Flask, g, render_template, request, abort, redirect, session, send_file

from openpyxl import Workbook, load_workbook

from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

app = Flask(__name__)
app.secret_key = "change-this-to-a-long-random-secret"

DB = "test.db"

# Admin password (per your request)
ADMIN_PASSWORD = "Rotamotion1"

# Certificate files (admin controlled)
CERT_DIR = "certificates"
CERT_XLSX = "certificates_log.xlsx"


# -------------------------
# Database helpers
# -------------------------
def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    conn = g.pop("db", None)
    if conn:
        conn.close()


def init_db():
    d = db()
    d.executescript("""
    PRAGMA foreign_keys = ON;

    CREATE TABLE IF NOT EXISTS tests (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      title TEXT NOT NULL,
      slug TEXT,
      pass_score INTEGER NOT NULL DEFAULT 70,
      created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS questions (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      test_id INTEGER NOT NULL,
      qtype TEXT NOT NULL DEFAULT 'MCQ' CHECK(qtype IN ('MCQ','TF')),
      prompt TEXT NOT NULL,
      a TEXT NOT NULL,
      b TEXT NOT NULL,
      c TEXT,
      d TEXT,
      correct CHAR(1) NOT NULL CHECK(correct IN ('A','B','C','D')),
      FOREIGN KEY(test_id) REFERENCES tests(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS attempts (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      test_id INTEGER NOT NULL,
      student_name TEXT NOT NULL,
      score INTEGER NOT NULL,
      passed INTEGER NOT NULL CHECK(passed IN (0,1)),
      created_at TEXT NOT NULL,
      FOREIGN KEY(test_id) REFERENCES tests(id) ON DELETE CASCADE
    );
    """)
    d.commit()

    # Migrations for older DBs
    test_cols = [r["name"] for r in d.execute("PRAGMA table_info(tests)").fetchall()]
    if "slug" not in test_cols:
        d.execute("ALTER TABLE tests ADD COLUMN slug TEXT")
        d.commit()

    q_cols = [r["name"] for r in d.execute("PRAGMA table_info(questions)").fetchall()]
    if "qtype" not in q_cols:
        d.execute("ALTER TABLE questions ADD COLUMN qtype TEXT NOT NULL DEFAULT 'MCQ'")
        d.commit()


def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    out = []
    prev_dash = False
    for ch in text:
        if ch.isalnum():
            out.append(ch)
            prev_dash = False
        else:
            if not prev_dash:
                out.append("-")
                prev_dash = True
    slug = "".join(out).strip("-")
    return slug or "test"


def unique_slug(base: str) -> str:
    s = base
    i = 2
    while db().execute("SELECT 1 FROM tests WHERE slug=?", (s,)).fetchone():
        s = f"{base}-{i}"
        i += 1
    return s


def ensure_slugs():
    tests = db().execute("SELECT id, title, slug FROM tests").fetchall()
    for t in tests:
        if not t["slug"]:
            slug = unique_slug(slugify(t["title"]))
            db().execute("UPDATE tests SET slug=? WHERE id=?", (slug, t["id"]))
    db().commit()


def seed_line_breaking_exam():
    title = "Line Breaking Final Exam"
    pass_score = 100

    existing = db().execute("SELECT id FROM tests WHERE title=?", (title,)).fetchone()
    if existing:
        return

    slug = unique_slug(slugify(title))
    cur = db().execute(
        "INSERT INTO tests(title, slug, pass_score, created_at) VALUES (?,?,?,?)",
        (title, slug, pass_score, datetime.utcnow().isoformat()),
    )
    test_id = cur.lastrowid

    questions = [
        ("MCQ",
         "Intentionally opening a pipe, line or duct for the purpose of cleaning, inspection, maintenance or replacing components within a system is referred to as:",
         "Line Breaking", "Pipe Sealing", "System Flushing", "Duct Isolation", "A"),

        ("MCQ",
         "Routine line breaking instructions can be found in the:",
         "Emergency action plan (EAP)", "PPE assessment (PPEA)", "Safety Data Sheets (SDS)", "Safe operation procedure (SOP)",
         "D"),

        ("MCQ",
         "Workers performing non-routine line-breaking tasks must obtain a(n) [ blank ] prior to starting work to identify the risks and controls that are needed.",
         "Safe Work Permit", "Safe operating Procedure (SOP)", "Emergency action plan (EAP)", "Open-end wrench", "A"),

        ("MCQ",
         "What type of flange slips over the pipe without needing to be welded to it and can swivel around the pipe to help line up opposing bolt holes?",
         "Orifice plate", "Lap joint flange", "Spectacle blind", "Blind flange", "B"),

        ("MCQ",
         "The minimum PPE for line-breaking jobs include a hard hat/helmet, gloves, face shield, goggles/safety glasses and:",
         "Hair net", "Chemical-protective clothing", "Disposable paper gown", "Shoe cover", "B"),

        ("MCQ",
         "To work in a defensive position opening a flange, begin by loosening the bolts [ blank. ]",
         "Farthest away from you.",
         "That are close to you first, working clockwise around the line.",
         "While crouched as low to the ground as possible.",
         "With open-end wrenches pointed toward your body.",
         "A"),

        ("TF",
         "When performing line breaking, it's best to use a cheater bar to gain leverage.",
         "True", "False", None, None, "B"),

        ("MCQ",
         "Before beginning line breaking, which of the following steps should be taken?",
         "Ensure lines and equipment are as free from recognized hazards as possible and test to confirm.",
         "Increase the pressure within the lines to check for leaks.",
         "Remove all tags from the equipment to allow for easy access.",
         "Turn on all cathodic protection rectifiers affecting the piping.",
         "A"),

        ("MCQ",
         "If the proper steps of bleeding off pressure are not followed, a flammable chemical release can create a potential ignition source, leading to:",
         "Biological hazards", "Fire and explosion.", "Slips, trips, and falls.", "Ultraviolet light exposure.",
         "B"),

        ("MCQ",
         "What should you do if you are unsure about which type of flange or procedure to use for line breaking?",
         "Check schematics and diagrams.", "Consult the human resources department.",
         "Perform the line break as best you can.", 'Use a "shop-made" flange.',
         "A"),
    ]

    for qtype, prompt, a, b, copt, dopt, correct in questions:
        # TF must store c/d as empty strings (not None) to avoid older NOT NULL DB issues
        if qtype == "TF":
            copt = ""
            dopt = ""
            correct = "A" if correct == "A" else "B"

        db().execute("""
            INSERT INTO questions(test_id, qtype, prompt, a, b, c, d, correct)
            VALUES (?,?,?,?,?,?,?,?)
        """, (test_id, qtype, prompt, a, b, copt, dopt, correct))

    db().commit()


@app.before_request
def _ensure_db():
    init_db()
    ensure_slugs()
    seed_line_breaking_exam()


# -------------------------
# Admin auth
# -------------------------
def is_admin() -> bool:
    return session.get("is_admin") is True


# -------------------------
# Certificate + Excel log
# -------------------------


def append_certificate_to_excel(test_title: str, student_name: str, score: int, pdf_path: str):
    os.makedirs(CERT_DIR, exist_ok=True)

    if os.path.exists(CERT_XLSX):
        wb = load_workbook(CERT_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Certificates"
        ws.append(["timestamp", "test_title", "student_name", "score", "pdf_file"])

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        test_title,
        student_name,
        score,
        pdf_path
    ])
    wb.save(CERT_XLSX)

def make_certificate_pdf(student_name: str, test_title: str, date_str: str) -> BytesIO:
    buf = BytesIO()

    pagesize = landscape(letter)
    c = canvas.Canvas(buf, pagesize=pagesize)
    width, height = pagesize

    c.setTitle("Certificate of Completion")

    # --- WATERMARK LOGO (faded, full-page cover) ---
    logo_path = os.path.join(app.root_path, "static", "logo.jpg")
    if os.path.exists(logo_path):
        try:
            img = ImageReader(logo_path)

            c.saveState()
            try:
                c.setFillAlpha(0.10)  # 0.06 lighter, 0.14 darker
            except Exception:
                pass

            margin = 0
            wm_w = width - (margin * 2)
            wm_h = height - (margin * 2)

            img_w, img_h = img.getSize()
            scale = max(wm_w / img_w, wm_h / img_h)
            draw_w = img_w * scale
            draw_h = img_h * scale

            x = margin + (wm_w - draw_w) / 2
            y = margin + (wm_h - draw_h) / 2

            c.drawImage(img, x, y, width=draw_w, height=draw_h, mask="auto")
            c.restoreState()
        except Exception:
            pass

    # --- TEXT ---
    c.setFont("Helvetica-Bold", 34)
    c.drawCentredString(width / 2, height - 110, "Certificate of Completion")

    c.setFont("Helvetica", 16)
    c.drawCentredString(width / 2, height - 165, "This certifies that")

    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(width / 2, height - 215, student_name)

    c.setFont("Helvetica", 16)
    c.drawCentredString(width / 2, height - 265, "has successfully completed")

    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(width / 2, height - 305, test_title)

    c.setFont("Helvetica", 14)
    c.drawCentredString(width / 2, 85, f"Date: {date_str}")

    # --- SIGNATURE (image) + name + line ---
    sig_path = os.path.join(app.root_path, "static", "signature.png")

    sig_x = 80
    sig_y = 70

    # Smaller signature
    sig_w = 200
    sig_h = 45

    if os.path.exists(sig_path):
        try:
            sig_img = ImageReader(sig_path)
            c.drawImage(sig_img, sig_x, sig_y, width=sig_w, height=sig_h, mask="auto")
        except Exception:
            # fallback line if image can't render
            c.setLineWidth(1)
            c.line(sig_x, sig_y + 15, sig_x + 200, sig_y + 15)
    else:
        # fallback line if file missing
        c.setLineWidth(1)
        c.line(sig_x, sig_y + 15, sig_x + 200, sig_y + 15)

    # Printed name under signature
    name_y = sig_y - 8
    c.setFont("Helvetica-Bold", 12)
    c.drawString(sig_x, name_y, "Chad Riley")

    # Line under the printed name
    line_y = name_y - 6
    c.setLineWidth(1)
    c.line(sig_x, line_y, sig_x + 180, line_y)

    # Label under the line
    c.setFont("Helvetica", 10)
    c.drawString(sig_x, line_y - 14, "Authorized Signature")

    c.showPage()
    c.save()

    buf.seek(0)
    return buf


def save_certificate_pdf(student_name: str, test_title: str) -> str:
    os.makedirs(CERT_DIR, exist_ok=True)

    date_str = datetime.now().strftime("%B %d, %Y")
    pdf_buf = make_certificate_pdf(student_name, test_title, date_str)

    safe_name = "".join(ch for ch in student_name if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Certificate_{safe_name}_{slugify(test_title)}_{timestamp}.pdf"
    pdf_path = os.path.join(CERT_DIR, filename)

    with open(pdf_path, "wb") as out:
        out.write(pdf_buf.getbuffer())

    return pdf_path


# -------------------------
# Admin: login/logout
# -------------------------
@app.route("/admin/login", methods=["GET"])
def admin_login():
    return """
    <html><body style="font-family:system-ui;max-width:420px;margin:40px auto;padding:0 16px;">
      <h1>Admin Login</h1>
      <form method="post" action="/admin/login">
        <input type="password" name="password" placeholder="Admin password"
               style="width:100%;padding:10px;margin:10px 0;" required />
        <button type="submit" style="width:100%;padding:10px;">Login</button>
      </form>
    </body></html>
    """


@app.route("/admin/login", methods=["POST"])
def admin_login_post():
    pw = (request.form.get("password") or "").strip()
    if pw == ADMIN_PASSWORD:
        session["is_admin"] = True
        return redirect("/admin/certificates")
    abort(403)


@app.route("/admin/logout", methods=["GET"])
def admin_logout():
    session.pop("is_admin", None)
    return redirect("/")


# -------------------------
# Admin: certificates dashboard (admin only)
# -------------------------
@app.route("/admin/certificates", methods=["GET"])
def admin_certificates():
    if not is_admin():
        return redirect("/admin/login")

    rows = []
    if os.path.exists(CERT_XLSX):
        wb = load_workbook(CERT_XLSX)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if len(data) >= 2:
            headers = data[0]
            for r in data[1:]:
                rows.append(dict(zip(headers, r)))

    html_rows = ""
    for r in reversed(rows):
        pdf_file = r.get("pdf_file") or ""
        pdf_name = os.path.basename(str(pdf_file)) if pdf_file else ""
        pdf_link = f'<a href="/admin/certificates/pdf/{pdf_name}">{pdf_name}</a>' if pdf_name else ""
        html_rows += f"""
          <tr>
            <td>{r.get("timestamp","")}</td>
            <td>{r.get("test_title","")}</td>
            <td>{r.get("student_name","")}</td>
            <td>{r.get("score","")}</td>
            <td>{pdf_link}</td>
          </tr>
        """

    return f"""
    <html>
      <head>
        <meta charset="utf-8" />
        <title>Admin Certificates</title>
        <style>
          body {{ font-family: system-ui, Arial; max-width: 1100px; margin: 32px auto; padding: 0 16px; }}
          a.btn {{ display:inline-block; padding:10px 14px; border:1px solid #444; border-radius:10px; text-decoration:none; margin-right: 8px; }}
          table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
          th, td {{ border-bottom: 1px solid #eee; padding: 10px; text-align: left; font-size: 14px; }}
          th {{ background: #fafafa; }}
          .muted {{ opacity: .75; }}
        </style>
      </head>
      <body>
        <h1>Certificates (Admin)</h1>
        <p class="muted">Only admins can view/download certificates and the Excel log.</p>

        <p>
          <a class="btn" href="/admin/certificates/excel">Download Excel Log</a>
          <a class="btn" href="/admin/logout">Logout</a>
          <a class="btn" href="/">Home</a>
        </p>

        <table>
          <thead>
            <tr>
              <th>Timestamp</th>
              <th>Test Title</th>
              <th>Student Name</th>
              <th>Score</th>
              <th>PDF</th>
            </tr>
          </thead>
          <tbody>
            {html_rows if html_rows else '<tr><td colspan="5" class="muted">No certificates logged yet.</td></tr>'}
          </tbody>
        </table>
      </body>
    </html>
    """


@app.route("/admin/certificates/excel", methods=["GET"])
def admin_download_excel():
    if not is_admin():
        return redirect("/admin/login")
    if not os.path.exists(CERT_XLSX):
        abort(404, "No Excel log yet.")
    return send_file(CERT_XLSX, as_attachment=True, download_name="certificates_log.xlsx")


@app.route("/admin/certificates/pdf/<path:filename>", methods=["GET"])
def admin_download_pdf(filename):
    if not is_admin():
        return redirect("/admin/login")

    safe = os.path.basename(filename)
    pdf_path = os.path.join(CERT_DIR, safe)

    if not os.path.exists(pdf_path):
        abort(404)

    return send_file(pdf_path, as_attachment=True, download_name=safe)


# -------------------------
# Home
# -------------------------
@app.route("/", methods=["GET"])
def home():
    tests = db().execute("SELECT * FROM tests ORDER BY id DESC").fetchall()
    return render_template("home.html", tests=tests)


# -------------------------
# Student: take test (by slug) + submit (by slug)
# -------------------------
@app.route("/tests/<slug>/take", methods=["GET"])
def take_test_by_slug(slug):
    t = db().execute("SELECT * FROM tests WHERE slug=?", (slug,)).fetchone()
    if not t:
        abort(404)

    qs = list(db().execute("SELECT * FROM questions WHERE test_id=?", (t["id"],)).fetchall())
    if not qs:
        abort(400, "This test has no questions yet.")

    random.shuffle(qs)
    saved_name = session.get("student_name", "")

    return render_template("take_test.html", test=t, questions=qs, saved_name=saved_name)


@app.route("/tests/<slug>/submit", methods=["POST"])
def submit_test_by_slug(slug):
    t = db().execute("SELECT * FROM tests WHERE slug=?", (slug,)).fetchone()
    if not t:
        abort(404)

    qs = list(db().execute("SELECT * FROM questions WHERE test_id=?", (t["id"],)).fetchall())
    if not qs:
        abort(400, "No questions yet")

    name = (request.form.get("student_name") or "").strip()
    if not name:
        abort(400, "Name required")

    session["student_name"] = name

    def opt_text(q, letter):
        mapping = {"A": q["a"], "B": q["b"], "C": q["c"], "D": q["d"]}
        val = mapping.get(letter)
        return val if val else ""

    review = []
    correct_count = 0

    for q in qs:
        chosen = (request.form.get(f"q_{q['id']}", "") or "").upper()
        is_correct = (chosen == q["correct"])
        if is_correct:
            correct_count += 1

        review.append({
            "prompt": q["prompt"],
            "qtype": q["qtype"],
            "chosen": chosen if chosen else "(No answer)",
            "chosen_text": opt_text(q, chosen) if chosen in ("A", "B", "C", "D") else "(No answer)",
            "correct": q["correct"],
            "correct_text": opt_text(q, q["correct"]),
            "is_correct": is_correct,
        })

    score = round((correct_count / len(qs)) * 100)
    passed = 1 if score >= t["pass_score"] else 0

    cur = db().execute("""
      INSERT INTO attempts(test_id, student_name, score, passed, created_at)
      VALUES (?,?,?,?,?)
    """, (t["id"], name, score, passed, datetime.utcnow().isoformat()))
    db().commit()
    attempt_id = cur.lastrowid

    # If passed: save admin copy + log to excel
    if passed == 1:
        pdf_path = save_certificate_pdf(name, t["title"])
        append_certificate_to_excel(t["title"], name, score, pdf_path)

    return render_template(
        "result.html",
        test=t,
        student_name=name,
        score=score,
        passed=bool(passed),
        review=review,
        attempt_id=attempt_id
    )


# Student download certificate (for their own attempt)
@app.route("/tests/<slug>/certificate/<int:attempt_id>", methods=["GET"])
def student_certificate(slug, attempt_id):
    t = db().execute("SELECT * FROM tests WHERE slug=?", (slug,)).fetchone()
    if not t:
        abort(404)

    a = db().execute("""
        SELECT * FROM attempts
        WHERE id=? AND test_id=?
    """, (attempt_id, t["id"])).fetchone()

    if not a:
        abort(404)

    if a["passed"] != 1:
        abort(403, "Certificate only available if you passed.")

    # Simple protection: only allow same browser session name
    session_name = session.get("student_name")
    if not session_name or session_name.strip() != a["student_name"]:
        abort(403, "This certificate is not available for your session.")

    date_str = datetime.now().strftime("%B %d, %Y")
    pdf_buf = make_certificate_pdf(a["student_name"], t["title"], date_str)

    safe_name = "".join(ch for ch in a["student_name"] if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
    filename = f"Certificate_{safe_name}_{slugify(t['title'])}.pdf"

    return send_file(pdf_buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


# Compatibility: old numeric routes redirect to slug
@app.route("/tests/<int:test_id>/take", methods=["GET"])
def take_test_numeric(test_id):
    t = db().execute("SELECT slug FROM tests WHERE id=?", (test_id,)).fetchone()
    if not t or not t["slug"]:
        abort(404)
    return redirect(f"/tests/{t['slug']}/take")


if __name__ == "__main__":
    app.run(debug=True)
